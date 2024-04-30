import argparse
import json
import re

import openpyxl

"""
rctc-to-json.py

Args:
python3 rctc-to-json.py spreadsheet_input_loc json_output_loc json_output_loc2

Arg list:
  1) spreadsheet_inpuc_loc: Input RCTC spreadsheet file location/name
  2) json_output_loc: output file location/name (Grouping List)
  3) json_output_loc2: output file location/name (Expansion List)

Example command:
python3 rctc-to-json.py reference-rctc-3-22.xlsx rctc_grouping rctc_expansion

The RCTC (Reportable Condition Trigger Codes) spreadsheet can be found
on the eRDS Aims Platform, and gets updated approximately every 6 months.

This Python script will read the RCTC spreadsheet and convert them to
machine-readable JSON files. See the Description markdown for more details.
"""


def convert_rctc_to_json(input_file, output_file_grouping, output_file_expansion):
    """
    Converts RCTC data from an Excel file to JSON, separating into 'grouping'
    and 'expansion'.

    :param input_file: Path to the input Excel file containing RCTC data.
    :param output_file_grouping: Path for the output JSON file with grouping data.
    :param output_file_expansion: Path for the output JSON file with expansion data.
    """
    wb = openpyxl.load_workbook(input_file)
    jsonData_grouping = {}
    jsonData_expansion = {}
    top_rows_buffer = 17  # number of rows before the Grouping data table begins
    tables_buffer = 7  # number of rows between Grouping & Expansion data table

    for sheet in wb.worksheets[2:]:
        ws = wb[sheet.title]
        title = sheet.title.replace("_", " ")
        title = re.sub(r" S\d", "", title)
        end_grouping = None

        # Grouping List
        for i, val in enumerate(
            ws.iter_rows(min_row=top_rows_buffer, values_only=True)
        ):
            if val[0] is None:
                end_grouping = i + top_rows_buffer
                break
            else:
                jsonData_grouping[val[1]] = {
                    "Type": title,
                    "Name": val[0],
                    "OID": val[1],
                    "Code System": val[2],
                    "Code System OID": val[3],
                    "Status": val[4],
                    "Condition Name": val[5],
                    "Condition Code": val[6],
                    "Condition Code System": val[7],
                }

        # Expansion List
        for val in ws.iter_rows(
            min_row=(end_grouping + tables_buffer), values_only=True
        ):
            if val[0] is None:
                break
            else:
                jsonData_expansion[val[1]] = {
                    "Type": title,
                    "Member OID": val[0],
                    "Code": val[1],
                    "Descriptor": val[2],
                    "Code System": val[3],
                    "Version": val[4],
                    "Status": val[5],
                    "Remap Info": val[6],
                }

    with (
        open(output_file_grouping, "w") as outfile1,
        open(output_file_expansion, "w") as outfile2,
    ):
        json.dump(jsonData_grouping, outfile1)

        json.dump(jsonData_expansion, outfile2)


def main():
    """
    Parses command line arguments and triggers RCTC spreadsheet conversion to JSON.
    """
    parser = argparse.ArgumentParser(
        description="Convert RCTC spreadsheet to JSON format."
    )
    parser.add_argument("input_file", help="Input RCTC spreadsheet location/name")
    parser.add_argument(
        "output_file_grouping", help="Output file location/name (Grouping List)"
    )
    parser.add_argument(
        "output_file_expansion", help="Output file location/name (Expansion List)"
    )
    args = parser.parse_args()

    convert_rctc_to_json(
        args.input_file, args.output_file_grouping, args.output_file_expansion
    )


if __name__ == "__main__":
    main()
